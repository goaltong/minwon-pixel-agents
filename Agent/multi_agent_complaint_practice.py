# Change log (created 2026-06-26):
# - Updated 2026-06-26 17:11 KST:
#   * Saved a backup before editing at OLD/multi_agent_complaint_practice_Ver.2606261711.py.
#   * Added OpenRouter retry and fallback-model handling so a temporary upstream
#     429 rate limit on the selected provider does not stop the whole workflow.
#   * Added --fallback-models, --llm-retries, and --llm-retry-delay options.
#   * Kept retry/fallback console messages in English.
# - Updated 2026-06-26 16:38 KST:
#   * Saved a backup before editing at OLD/multi_agent_complaint_practice_Ver.2606261638.py.
#   * Added optional Pixel Agents hook reporting for the OpenRouter/local Python
#     workflow without requiring Claude Code.
#   * Added --pixel-agents, --pixel-agents-provider, and --pixel-agents-session-id
#     command-line options.
#   * Emits session, stage start/end, progress, and completion events to the
#     running Pixel Agents server discovered from ~/.pixel-agents/server.json.
#   * Kept event delivery best-effort so API/model/output failures remain visible
#     through the normal pipeline exceptions instead of being hidden by telemetry.
#   * Ensured Pixel Agents receives an error session-end event if the pipeline
#     raises during a stage.
# - Updated 2026-06-26 14:56 KST:
#   * Saved a backup before editing at OLD/multi_agent_complaint_practice_Ver.2606261456.py.
#   * Localized classification departments, complaint types, legal/policy basis names,
#     and basis summaries to Korean so the web UI and result files no longer mix
#     English labels into Korean petition drafts.
# - Updated 2026-06-26 14:55 KST:
#   * Saved a backup before editing at OLD/multi_agent_complaint_practice_Ver.2606261455.py.
#   * Changed local template draft responses from English to Korean for Korean civil
#     complaint practice.
#   * Changed the LLM draft prompt so generated petition responses are written in Korean.
#   * Updated ReviewAgent to accept Korean required phrases while still accepting
#     older English drafts for compatibility.
#   * Updated Markdown report labels from English to Korean. Console messages remain
#     English as required by the project instruction.
# - Updated 2026-06-26 14:44 KST:
#   * Saved a backup before editing at OLD/multi_agent_complaint_practice_Ver.2606261444.py.
#   * Tightened the LLM draft prompt so the model must quote the exact legal/policy
#     basis name supplied by SearchAgent.
#   * Made ReviewAgent's basis check case-insensitive while still requiring the
#     exact basis phrase to be present in the draft.
# - Updated 2026-06-26 14:43 KST:
#   * Saved a backup before editing at OLD/multi_agent_complaint_practice_Ver.2606261443.py.
#   * Added optional LLM mode using the API key stored in Key.txt.
#   * Added an OpenRouter-compatible chat-completions client implemented with Python
#     standard library urllib to avoid extra SDK installation.
#   * Added --list-models, --model, --api-key-file, --api-base-url, and --use-llm
#     command-line options.
#   * Added the model names shown in the provided folder image as selectable model
#     aliases, while keeping local rule/template mode as the default to avoid
#     accidental API cost.
# - Updated 2026-06-26 14:41 KST:
#   * Saved a backup before editing at OLD/multi_agent_complaint_practice_Ver.2606261441.py.
#   * Changed Markdown output encoding from plain UTF-8 to UTF-8 with BOM so Korean
#     complaint titles and bodies display correctly in Windows PowerShell/Get-Content.
# - Added a low-cost, local-only practice implementation for the complaint drafting
#   multi-agent workflow shown in the lesson slide.
# - The workflow is split into four explicit handoff stages:
#   1) ClassificationAgent: classifies each complaint by likely department/type.
#   2) SearchAgent: attaches likely legal/policy basis using a small local rule base.
#   3) DraftAgent: creates a structured draft response from complaint facts and basis.
#   4) ReviewAgent: checks whether the draft contains key elements and flags issues.
# - The script reads the provided Excel workbook without modifying it.
# - The script writes English console output and English output files to keep runtime
#   outputs consistent with the project instruction.
# - This implementation intentionally avoids paid LLM/API calls. To extend it later,
#   replace SearchAgent and DraftAgent internals with real retrieval/model calls while
#   keeping the same handoff payload shape.

from __future__ import annotations

import argparse
import csv
import html
import json
import os
import re
import time
import urllib.error
import urllib.request
import uuid
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


DEFAULT_INPUT = Path("1782341096215_file (2).xlsx")
DEFAULT_OUTPUT_DIR = Path("outputs")
DEFAULT_API_KEY_FILE = Path("Key.txt")
DEFAULT_API_BASE_URL = "https://openrouter.ai/api/v1"
DEFAULT_MODEL = "microsoft/phi-4"
DEFAULT_FALLBACK_MODELS = ["google/gemini-2.5-flash", "mistralai/mistral-small-3.2-24b-instruct"]
PIXEL_AGENTS_SERVER_JSON = Path.home() / ".pixel-agents" / "server.json"

AVAILABLE_MODELS = {
    "nvidia": "nvidia/nemotron-3-super-120b-a12b",
    "gemma": "google/gemma-4-31b-it",
    "phi": "microsoft/phi-4",
    "llama": "meta-llama/llama-4-maverick",
    "mistral": "mistralai/mistral-small-3.2-24b-instruct",
    "qwen": "qwen/qwen3.5-plus-02-15",
    "gemini": "google/gemini-2.5-flash",
}


@dataclass
class Complaint:
    complaint_id: str
    title: str
    body: str


@dataclass
class ClassificationResult:
    complaint_id: str
    title: str
    department: str
    complaint_type: str
    confidence: float
    matched_keywords: str


@dataclass
class SearchResult:
    complaint_id: str
    legal_basis: str
    basis_summary: str
    search_notes: str


@dataclass
class DraftResult:
    complaint_id: str
    draft_response: str


@dataclass
class ReviewResult:
    complaint_id: str
    review_status: str
    review_notes: str


@dataclass
class PipelineResult:
    complaint_id: str
    title: str
    department: str
    complaint_type: str
    confidence: float
    matched_keywords: str
    legal_basis: str
    basis_summary: str
    draft_response: str
    review_status: str
    review_notes: str


class ClassificationAgent:
    RULES = [
        (
            "채용 및 시험",
            "응시 자격 또는 시험 문의",
            ["7급", "응시", "시험", "국가고시", "한국사", "채용", "지방인재"],
        ),
        (
            "인사 관리",
            "임용, 승진 또는 휴직 문의",
            ["임용", "승진", "휴직", "육아휴직", "인사", "복무"],
        ),
        (
            "출장 여비",
            "출장비 지급 기준 문의",
            ["출장", "자가차량", "유가", "연비", "여비"],
        ),
        (
            "간행물 및 정보공개",
            "간행물 또는 자료 제공 문의",
            ["간행물", "자료", "책자", "발간", "배포"],
        ),
        (
            "디지털 서비스 지원",
            "계정 또는 웹사이트 이용 문의",
            ["비밀번호", "로그인", "이메일", "사이버국가고시센터", "인증"],
        ),
    ]

    def run(self, complaint: Complaint) -> ClassificationResult:
        text = f"{complaint.title}\n{complaint.body}"
        best = ("일반 민원", "일반 문의", [], 0)
        for department, complaint_type, keywords in self.RULES:
            matched = [keyword for keyword in keywords if keyword in text]
            if len(matched) > best[3]:
                best = (department, complaint_type, matched, len(matched))

        confidence = min(0.95, 0.45 + best[3] * 0.12)
        return ClassificationResult(
            complaint_id=complaint.complaint_id,
            title=complaint.title,
            department=best[0],
            complaint_type=best[1],
            confidence=round(confidence, 2),
            matched_keywords=", ".join(best[2]) if best[2] else "No strong keyword match",
        )


class SearchAgent:
    BASIS = {
        "채용 및 시험": (
            "공무원 채용시험 공고 및 응시자격 기준",
            "최신 시험 공고, 자격 인정 기준, 응시자 안내사항을 확인해야 합니다.",
        ),
        "인사 관리": (
            "국가공무원법 및 공무원임용령",
            "임용, 승진소요기간, 휴직, 복무상 지위 관련 조항을 확인해야 합니다.",
        ),
        "출장 여비": (
            "공무원 여비 규정",
            "자가차량 이용, 이동거리, 유류비, 지급 산식이 여비 지급 기준에 부합하는지 확인해야 합니다.",
        ),
        "간행물 및 정보공개": (
            "공공기관의 정보공개에 관한 법률 및 간행물 배포 기준",
            "간행물 소관 부서, 배포 경로, 재고 여부, 국민 열람 또는 구매 가능 경로를 확인해야 합니다.",
        ),
        "디지털 서비스 지원": (
            "전자정부서비스 운영 지침",
            "본인 확인, 이메일 발송, 계정 복구, 고객지원 절차를 확인해야 합니다.",
        ),
        "일반 민원": (
            "소관 기관 민원 처리 기준",
            "담당 부서를 확인하고, 제출된 사실관계가 부족한 경우 보완 자료를 요청해야 합니다.",
        ),
    }

    def run(self, classification: ClassificationResult) -> SearchResult:
        legal_basis, summary = self.BASIS[classification.department]
        return SearchResult(
            complaint_id=classification.complaint_id,
            legal_basis=legal_basis,
            basis_summary=summary,
            search_notes="Local rule-base search used. Replace with web/RAG search for production.",
        )


class ChatClient:
    def __init__(
        self,
        api_key: str,
        model: str,
        api_base_url: str,
        fallback_models: list[str] | None = None,
        max_retries: int = 2,
        retry_delay_seconds: float = 2.0,
    ) -> None:
        self.api_key = api_key
        self.model = model
        self.api_base_url = api_base_url.rstrip("/")
        self.fallback_models = fallback_models or []
        self.max_retries = max(0, max_retries)
        self.retry_delay_seconds = max(0.0, retry_delay_seconds)

    def complete(self, system_prompt: str, user_prompt: str) -> str:
        models = [self.model, *[model for model in self.fallback_models if model != self.model]]
        last_error: RuntimeError | None = None
        for model_index, model in enumerate(models):
            for attempt in range(self.max_retries + 1):
                try:
                    return self._complete_with_model(model, system_prompt, user_prompt)
                except RuntimeError as exc:
                    last_error = exc
                    if not self._is_retryable_error(exc):
                        raise
                    has_retry = attempt < self.max_retries
                    has_next_model = model_index < len(models) - 1
                    if has_retry:
                        wait_seconds = self.retry_delay_seconds * (attempt + 1)
                        print(
                            f"LLM request retryable failure for {model}; retrying in {wait_seconds:.1f}s."
                        )
                        time.sleep(wait_seconds)
                        continue
                    if has_next_model:
                        next_model = models[model_index + 1]
                        print(
                            f"LLM model {model} is unavailable or rate-limited; trying fallback {next_model}."
                        )
                        break
                    raise
        if last_error:
            raise last_error
        raise RuntimeError("LLM request failed: no model configured.")

    def _complete_with_model(self, model: str, system_prompt: str, user_prompt: str) -> str:
        payload = {
            "model": model,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            "temperature": 0.2,
            "max_tokens": 700,
        }
        data = json.dumps(payload).encode("utf-8")
        request = urllib.request.Request(
            f"{self.api_base_url}/chat/completions",
            data=data,
            headers={
                "Authorization": f"Bearer {self.api_key}",
                "Content-Type": "application/json",
                "HTTP-Referer": "http://localhost/multi-agent-practice",
                "X-Title": "Complaint Multi-Agent Practice",
            },
            method="POST",
        )
        try:
            with urllib.request.urlopen(request, timeout=60) as response:
                response_data = json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as exc:
            error_body = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"LLM request failed: HTTP {exc.code} {error_body}") from exc
        except urllib.error.URLError as exc:
            raise RuntimeError(f"LLM request failed: {exc.reason}") from exc

        try:
            return response_data["choices"][0]["message"]["content"].strip()
        except (KeyError, IndexError, TypeError) as exc:
            raise RuntimeError(f"Unexpected LLM response shape: {response_data}") from exc

    @staticmethod
    def _is_retryable_error(exc: RuntimeError) -> bool:
        message = str(exc)
        return (
            "HTTP 429" in message
            or "HTTP 500" in message
            or "HTTP 502" in message
            or "HTTP 503" in message
            or "HTTP 504" in message
            or "timed out" in message.lower()
            or "temporarily" in message.lower()
        )


class PixelAgentsReporter:
    def __init__(
        self,
        enabled: bool,
        provider_id: str,
        session_id: str | None = None,
        cwd: Path | None = None,
    ) -> None:
        self.enabled = enabled
        self.provider_id = provider_id
        self.session_id = session_id or str(uuid.uuid4())
        self.cwd = str((cwd or Path.cwd()).resolve())
        self.server: dict[str, object] | None = None
        self.current_tool_id: str | None = None
        if self.enabled:
            self.server = self._read_server_config()

    def _read_server_config(self) -> dict[str, object] | None:
        try:
            return json.loads(PIXEL_AGENTS_SERVER_JSON.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            print(f"Pixel Agents reporting disabled: {exc}")
            return None

    def emit(self, event: dict[str, object]) -> None:
        if not self.enabled or not self.server:
            return
        port = self.server.get("port")
        token = self.server.get("token")
        if not isinstance(port, int) or not isinstance(token, str):
            return
        payload = {
            "session_id": self.session_id,
            "cwd": self.cwd,
            **event,
        }
        data = json.dumps(payload).encode("utf-8")
        request = urllib.request.Request(
            f"http://127.0.0.1:{port}/api/hooks/{self.provider_id}",
            data=data,
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
            },
            method="POST",
        )
        try:
            with urllib.request.urlopen(request, timeout=2):
                pass
        except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError) as exc:
            print(f"Pixel Agents event delivery failed: {exc}")

    def session_start(self) -> None:
        self.emit({"hook_event_name": "SessionStart", "source": "local-agent"})

    def session_end(self, reason: str = "completed") -> None:
        self.emit({"hook_event_name": "SessionEnd", "reason": reason})

    def stage_start(self, stage_name: str, complaint_id: str) -> str:
        tool_id = f"{stage_name}-{complaint_id}-{time.time_ns()}"
        self.current_tool_id = tool_id
        self.emit(
            {
                "hook_event_name": "PreToolUse",
                "tool_id": tool_id,
                "tool_name": stage_name,
                "tool_input": {"complaint_id": complaint_id},
            }
        )
        return tool_id

    def stage_end(self, tool_id: str) -> None:
        self.emit({"hook_event_name": "PostToolUse", "tool_id": tool_id})
        if self.current_tool_id == tool_id:
            self.current_tool_id = None

    def progress(self, tool_id: str, message: str) -> None:
        self.emit({"hook_event_name": "Progress", "tool_id": tool_id, "data": {"message": message}})

    def turn_end(self) -> None:
        self.emit({"hook_event_name": "Stop"})


class NullPixelAgentsReporter(PixelAgentsReporter):
    def __init__(self) -> None:
        self.enabled = False
        self.provider_id = "local-agent"
        self.session_id = ""
        self.cwd = ""
        self.server = None
        self.current_tool_id = None

    def emit(self, event: dict[str, object]) -> None:
        return


class DraftAgent:
    def __init__(self, chat_client: ChatClient | None = None) -> None:
        self.chat_client = chat_client

    def run(
        self,
        complaint: Complaint,
        classification: ClassificationResult,
        search: SearchResult,
    ) -> DraftResult:
        if self.chat_client:
            return self._run_with_llm(complaint, classification, search)

        body_summary = summarize_text(complaint.body, max_chars=260)
        draft = (
            "안녕하십니까.\n\n"
            "민원을 제출해 주셔서 감사합니다.\n\n"
            f"귀하께서 문의하신 내용은 '{complaint.title}'에 관한 사항으로 이해됩니다.\n"
            f"제출하신 민원 요지는 다음과 같습니다: {body_summary}\n\n"
            f"이 민원은 '{classification.department}' 분야의 "
            f"'{classification.complaint_type}' 유형으로 분류됩니다.\n"
            f"검토할 관련 근거: {search.legal_basis}\n"
            f"근거 요약: {search.basis_summary}\n\n"
            "제출된 내용만으로는 최종 판단을 확정하기 어려우므로, 담당자는 사실관계와 최신 기준을 확인한 뒤 "
            "적용 가능한 절차, 필요 서류, 처리 기한 또는 담당 연락처를 포함하여 최종 답변을 작성해야 합니다.\n\n"
            "본 문안은 자동 생성된 답변 초안이며, 발송 전 반드시 담당자의 최종 검토가 필요합니다."
        )
        return DraftResult(complaint_id=complaint.complaint_id, draft_response=draft)

    def _run_with_llm(
        self,
        complaint: Complaint,
        classification: ClassificationResult,
        search: SearchResult,
    ) -> DraftResult:
        system_prompt = (
            "당신은 민원 답변 초안을 작성하는 에이전트입니다. "
            "답변은 반드시 자연스러운 한국어로 작성하십시오. "
            "법령명이나 최종 판단을 지어내지 말고, 사람 검토가 필요하다는 문구를 명확히 포함하십시오."
        )
        user_prompt = (
            f"Complaint ID: {complaint.complaint_id}\n"
            f"Title: {complaint.title}\n"
            f"Body: {summarize_text(complaint.body, max_chars=1200)}\n"
            f"Classified department: {classification.department}\n"
            f"Complaint type: {classification.complaint_type}\n"
            f"Likely basis: {search.legal_basis}\n"
            f"Basis summary: {search.basis_summary}\n\n"
            f"You must include this exact basis phrase in the draft: {search.legal_basis}\n\n"
            "아래 항목을 포함해 한국어 답변 초안을 작성하십시오.\n"
            "1. 인사\n"
            "2. 민원 내용 이해\n"
            "3. 검토할 관련 근거. 위의 exact basis phrase는 원문 그대로 포함하십시오.\n"
            "4. 답변 초안\n"
            "5. 담당자 최종 검토 필요 문구"
        )
        return DraftResult(
            complaint_id=complaint.complaint_id,
            draft_response=self.chat_client.complete(system_prompt, user_prompt),
        )


class ReviewAgent:
    REQUIRED_TERM_GROUPS = [
        ["감사", "thank you"],
        ["관련 근거", "relevant basis"],
        ["검토", "human review"],
    ]

    def run(self, draft: DraftResult, search: SearchResult) -> ReviewResult:
        draft_lower = draft.draft_response.lower()
        missing = [
            " / ".join(group)
            for group in self.REQUIRED_TERM_GROUPS
            if not any(term in draft_lower for term in group)
        ]
        if search.legal_basis.lower() not in draft_lower:
            missing.append("근거명")

        if missing:
            return ReviewResult(
                complaint_id=draft.complaint_id,
                review_status="수정 필요",
                review_notes="누락된 항목: " + ", ".join(missing),
            )

        return ReviewResult(
            complaint_id=draft.complaint_id,
            review_status="통과",
            review_notes="인사, 민원 요지, 관련 근거, 담당자 검토 문구가 포함되어 있습니다.",
        )


def summarize_text(text: str, max_chars: int) -> str:
    normalized = re.sub(r"\s+", " ", text).strip()
    if len(normalized) <= max_chars:
        return normalized
    return normalized[: max_chars - 3].rstrip() + "..."


def read_complaints(path: Path, limit: int) -> list[Complaint]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    header_index = {name: idx for idx, name in enumerate(headers)}

    required = ["민원신청번호", "제목", "본문"]
    missing = [name for name in required if name not in header_index]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    complaints: list[Complaint] = []
    for row in rows:
        complaint_id = clean_cell(row[header_index["민원신청번호"]])
        title = clean_cell(row[header_index["제목"]])
        body = clean_cell(row[header_index["본문"]])
        if not complaint_id or not title:
            continue
        complaints.append(Complaint(complaint_id=complaint_id, title=title, body=body))
        if len(complaints) >= limit:
            break

    return complaints


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    return html.unescape(str(value)).strip()


def run_pipeline(
    complaints: Iterable[Complaint],
    chat_client: ChatClient | None = None,
    pixel_agents: PixelAgentsReporter | None = None,
) -> list[PipelineResult]:
    classifier = ClassificationAgent()
    searcher = SearchAgent()
    drafter = DraftAgent(chat_client=chat_client)
    reviewer = ReviewAgent()
    reporter = pixel_agents or NullPixelAgentsReporter()

    results: list[PipelineResult] = []
    reporter.session_start()
    try:
        for complaint in complaints:
            tool_id = reporter.stage_start("ClassificationAgent", complaint.complaint_id)
            classification = classifier.run(complaint)
            reporter.progress(tool_id, "Classification completed.")
            reporter.stage_end(tool_id)

            tool_id = reporter.stage_start("SearchAgent", complaint.complaint_id)
            search = searcher.run(classification)
            reporter.progress(tool_id, "Policy basis search completed.")
            reporter.stage_end(tool_id)

            tool_id = reporter.stage_start("DraftAgent", complaint.complaint_id)
            draft = drafter.run(complaint, classification, search)
            reporter.progress(tool_id, "Draft response completed.")
            reporter.stage_end(tool_id)

            tool_id = reporter.stage_start("ReviewAgent", complaint.complaint_id)
            review = reviewer.run(draft, search)
            reporter.progress(tool_id, "Review completed.")
            reporter.stage_end(tool_id)

            results.append(
                PipelineResult(
                    complaint_id=complaint.complaint_id,
                    title=complaint.title,
                    department=classification.department,
                    complaint_type=classification.complaint_type,
                    confidence=classification.confidence,
                    matched_keywords=classification.matched_keywords,
                    legal_basis=search.legal_basis,
                    basis_summary=search.basis_summary,
                    draft_response=draft.draft_response,
                    review_status=review.review_status,
                    review_notes=review.review_notes,
                )
            )
    except Exception:
        reporter.session_end("error")
        raise
    reporter.turn_end()
    reporter.session_end()
    return results


def write_csv(results: list[PipelineResult], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=list(asdict(results[0]).keys()))
        writer.writeheader()
        for result in results:
            writer.writerow(asdict(result))


def write_markdown(results: list[PipelineResult], path: Path) -> None:
    lines = ["# 민원 Multi-Agent 실습 결과", ""]
    for result in results:
        lines.extend(
            [
                f"## {result.complaint_id} - {result.title}",
                "",
                f"- 담당 영역: {result.department}",
                f"- 민원 유형: {result.complaint_type}",
                f"- 신뢰도: {result.confidence}",
                f"- 관련 근거: {result.legal_basis}",
                f"- 검수 결과: {result.review_status}",
                "",
                "### 답변 초안",
                "",
                result.draft_response,
                "",
                "### 검수 의견",
                "",
                result.review_notes,
                "",
            ]
        )
    path.write_text("\n".join(lines), encoding="utf-8-sig")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run a local multi-agent complaint drafting practice workflow."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Input Excel file path.")
    parser.add_argument("--limit", type=int, default=10, help="Number of complaints to process.")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="Directory for generated result files.",
    )
    parser.add_argument(
        "--use-llm",
        action="store_true",
        help="Use an OpenRouter-compatible LLM for the draft stage.",
    )
    parser.add_argument(
        "--pixel-agents",
        action="store_true",
        help="Report pipeline stage events to a running Pixel Agents server.",
    )
    parser.add_argument(
        "--pixel-agents-provider",
        default="local-agent",
        help="Pixel Agents provider id for hook reporting.",
    )
    parser.add_argument(
        "--pixel-agents-session-id",
        default=None,
        help="Optional fixed Pixel Agents session id.",
    )
    parser.add_argument(
        "--model",
        default=DEFAULT_MODEL,
        help="Model name or alias. Use --list-models to see aliases.",
    )
    parser.add_argument(
        "--fallback-models",
        default=",".join(DEFAULT_FALLBACK_MODELS),
        help="Comma-separated model names or aliases to try after retryable LLM failures.",
    )
    parser.add_argument(
        "--llm-retries",
        type=int,
        default=1,
        help="Retry count per model for retryable LLM failures.",
    )
    parser.add_argument(
        "--llm-retry-delay",
        type=float,
        default=2.0,
        help="Base delay in seconds before retrying a retryable LLM failure.",
    )
    parser.add_argument(
        "--api-key-file",
        type=Path,
        default=DEFAULT_API_KEY_FILE,
        help="Text file containing the API key.",
    )
    parser.add_argument(
        "--api-base-url",
        default=DEFAULT_API_BASE_URL,
        help="OpenAI-compatible API base URL.",
    )
    parser.add_argument(
        "--list-models",
        action="store_true",
        help="Print available model aliases and exit.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.list_models:
        print("Available model aliases:")
        for alias, model_name in AVAILABLE_MODELS.items():
            default_mark = " (default)" if model_name == DEFAULT_MODEL else ""
            print(f"- {alias}: {model_name}{default_mark}")
        print("You can also pass a full model name directly with --model.")
        return

    model = resolve_model(args.model)
    fallback_models = resolve_model_list(args.fallback_models)
    chat_client = None
    if args.use_llm:
        api_key = read_api_key(args.api_key_file)
        chat_client = ChatClient(
            api_key=api_key,
            model=model,
            api_base_url=args.api_base_url,
            fallback_models=fallback_models,
            max_retries=args.llm_retries,
            retry_delay_seconds=args.llm_retry_delay,
        )

    complaints = read_complaints(args.input, args.limit)
    if not complaints:
        raise SystemExit("No complaints found in the input workbook.")

    reporter = PixelAgentsReporter(
        enabled=args.pixel_agents,
        provider_id=args.pixel_agents_provider,
        session_id=args.pixel_agents_session_id,
        cwd=Path(os.getcwd()),
    )
    results = run_pipeline(complaints, chat_client=chat_client, pixel_agents=reporter)
    csv_path = args.output_dir / "complaint_multi_agent_results.csv"
    markdown_path = args.output_dir / "complaint_multi_agent_results.md"
    write_csv(results, csv_path)
    write_markdown(results, markdown_path)

    print("Multi-agent complaint workflow completed.")
    print(f"Input file: {args.input}")
    print(f"Processed complaints: {len(results)}")
    print(f"Draft mode: {'LLM' if args.use_llm else 'Local template'}")
    print(f"Model: {model if args.use_llm else 'not used'}")
    if args.use_llm and fallback_models:
        print(f"Fallback models: {', '.join(fallback_models)}")
    print(f"CSV output: {csv_path}")
    print(f"Markdown output: {markdown_path}")
    print("Stage handoff: classify -> search -> draft -> review -> human")


def resolve_model(model_or_alias: str) -> str:
    return AVAILABLE_MODELS.get(model_or_alias, model_or_alias)


def resolve_model_list(models: str) -> list[str]:
    return [resolve_model(model.strip()) for model in models.split(",") if model.strip()]


def read_api_key(path: Path) -> str:
    if not path.exists():
        raise FileNotFoundError(f"API key file not found: {path}")
    api_key = path.read_text(encoding="utf-8", errors="replace").strip()
    if not api_key:
        raise ValueError(f"API key file is empty: {path}")
    return api_key


if __name__ == "__main__":
    main()
